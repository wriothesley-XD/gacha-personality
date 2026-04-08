"""
convert.py — Konversi template_karakter.xlsx → characters.json
Cara pakai: python convert.py
Atau dengan nama file custom: python convert.py nama_file.xlsx output.json
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl belum terinstall. Jalankan: pip install openpyxl")
    sys.exit(1)

# ── Konfigurasi file ────────────────────────────────────────────────
EXCEL_FILE  = sys.argv[1] if len(sys.argv) > 1 else "template_karakter.xlsx"
OUTPUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "characters.json"
SHEET_NAME  = "Karakter"
DATA_START_ROW = 5

# ── Kolom mapping ───────────────────────────────────────────────────
COL = {
    "name": 1, "gender": 2, "img": 3, "sound": 4,
    "desc": 5, "details": 6, "words": 7,
    "brave": 8, "smart": 9, "gentle": 10,
    "leader": 11, "warm": 12, "cautious": 13,
    "mapping": 14,
}

VALID_GENDERS = {"male", "female", "anomali"}

def clean(val):
    return "" if val is None else str(val).strip()

def clamp_trait(val):
    try:
        return max(0, min(3, int(float(str(val)))))
    except (ValueError, TypeError):
        return 0

def convert():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ File tidak ditemukan: {EXCEL_FILE}")
        sys.exit(1)

    print(f"📂 Membaca: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' tidak ditemukan. Sheet yang ada: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    characters = []
    name_mappings = {}
    errors = []
    skipped = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        def get(col_key):
            return clean(ws.cell(row=row_idx, column=COL[col_key]).value)

        name   = get("name")
        gender = get("gender").lower()

        if not name:
            skipped += 1
            continue

        if gender not in VALID_GENDERS:
            errors.append(f"  Baris {row_idx}: gender '{gender}' tidak valid — karakter '{name}' dilewati")
            continue

        char = {
            "name":    name,
            "gender":  gender,
            "img":     get("img") or f"images/{name.lower().replace(' ', '_')}.jpg",
            "sound":   get("sound") or "",
            "desc":    get("desc") or "",
            "details": get("details") or "",
            "words":   get("words") or "",
            "traits": {
                "brave":    clamp_trait(get("brave")),
                "smart":    clamp_trait(get("smart")),
                "gentle":   clamp_trait(get("gentle")),
                "leader":   clamp_trait(get("leader")),
                "warm":     clamp_trait(get("warm")),
                "cautious": clamp_trait(get("cautious")),
            }
        }
        characters.append(char)

        mapping = get("mapping")
        if mapping:
            for raw in mapping.split(","):
                raw = raw.strip()
                if raw:
                    name_mappings[raw.lower()] = name

    output = {"characters": characters, "nameMapping": name_mappings}

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Selesai!")
    print(f"   Karakter berhasil  : {len(characters)}")
    print(f"   Mapping nama       : {len(name_mappings)}")
    print(f"   Baris dilewati     : {skipped} (kosong)")

    if errors:
        print(f"\n⚠️  {len(errors)} baris ada masalah:")
        for e in errors:
            print(e)

    print(f"\n📄 Output: {OUTPUT_FILE}")
    print("   Letakkan file ini di folder yang sama dengan index.html")

if __name__ == "__main__":
    convert()
